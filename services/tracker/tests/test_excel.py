"""Tests des Fortschreibens der Tabelle.

Der Kern ist nicht das Schreiben, sondern das Zusammenfuehren: ein
zweiter Export darf die eigenen Eintragungen nicht ueberbuegeln.
"""
import pytest
from openpyxl import Workbook, load_workbook

from tracker import excel, felder, status


def zeile(referenz="10001-1-S", firma="Beispiel GmbH", **rest):
    werte = {
        felder.SPALTE_REFERENZ: referenz,
        "Firma": firma,
        "Position": "Data Engineer",
        "Link zur Ausschreibung": f"{felder.STELLENLINK}{referenz}",
        "Standort": "20095 Hamburg",
        "Homeoffice-Modell": "100 % remote",
        "Gehalt": "",
        "Benefits": "Jobticket / ÖPNV",
        "Kontakt": "Frau Meyer · a@b.de",
        "Status": "",
    }
    werte.update(rest)
    return werte


def lies(pfad):
    blatt = load_workbook(pfad).active
    kopf = [z.value for z in blatt[1]]
    return blatt, {name: nummer for nummer, name in enumerate(kopf, start=1) if name}


def test_neue_datei_bekommt_alle_spalten(tmp_path):
    ziel = tmp_path / "tracker.xlsx"

    bericht = excel.schreibe(ziel, [zeile()])

    assert bericht.neu == 1
    assert bericht.sicherung is None
    blatt, spalten = lies(ziel)
    for name in felder.SPALTEN:
        assert name in spalten
    assert felder.SPALTE_REFERENZ in spalten
    assert blatt.cell(row=2, column=spalten["Firma"]).value == "Beispiel GmbH"
    assert blatt.cell(row=2, column=spalten["Nr."]).value == 1


def test_zweiter_lauf_aktualisiert_statt_anzuhaengen(tmp_path):
    ziel = tmp_path / "tracker.xlsx"
    excel.schreibe(ziel, [zeile()])

    bericht = excel.schreibe(ziel, [zeile(firma="Umbenannt AG")])

    assert (bericht.neu, bericht.aktualisiert) == (0, 1)
    blatt, spalten = lies(ziel)
    assert blatt.max_row == 2
    assert blatt.cell(row=2, column=spalten["Firma"]).value == "Umbenannt AG"


def test_handgepflegte_spalten_bleiben_unangetastet(tmp_path):
    ziel = tmp_path / "tracker.xlsx"
    excel.schreibe(ziel, [zeile()])

    mappe = load_workbook(ziel)
    blatt = mappe.active
    _, spalten = lies(ziel)
    blatt.cell(row=2, column=spalten["Kontakt"], value="Telefonat mit Frau Meyer")
    blatt.cell(row=2, column=spalten["Gehalt"], value="65.000 EUR, verhandelt")
    mappe.save(ziel)

    excel.schreibe(ziel, [zeile()])

    blatt, spalten = lies(ziel)
    # Eine eigene Eintragung wiegt schwerer als ein Vorschlag.
    assert blatt.cell(row=2, column=spalten["Kontakt"]).value == "Telefonat mit Frau Meyer"
    assert blatt.cell(row=2, column=spalten["Gehalt"]).value == "65.000 EUR, verhandelt"


def test_ueberschreiben_erneuert_auch_die_vorschlaege(tmp_path):
    ziel = tmp_path / "tracker.xlsx"
    excel.schreibe(ziel, [zeile()])
    mappe = load_workbook(ziel)
    _, spalten = lies(ziel)
    mappe.active.cell(row=2, column=spalten["Benefits"], value="alt")
    mappe.save(ziel)

    excel.schreibe(ziel, [zeile()], ueberschreiben=True)

    blatt, spalten = lies(ziel)
    assert blatt.cell(row=2, column=spalten["Benefits"]).value == "Jobticket / ÖPNV"


def test_leerer_wert_loescht_nichts_vorhandenes(tmp_path):
    """Eine ausgefallene Quelle darf keine Daten vernichten."""
    ziel = tmp_path / "tracker.xlsx"
    excel.schreibe(ziel, [zeile()])

    excel.schreibe(ziel, [zeile(**{"Standort": "", "Firma": ""})])

    blatt, spalten = lies(ziel)
    assert blatt.cell(row=2, column=spalten["Standort"]).value == "20095 Hamburg"
    assert blatt.cell(row=2, column=spalten["Firma"]).value == "Beispiel GmbH"


def test_bestehende_zeile_wird_ueber_den_link_erkannt(tmp_path):
    """Der erste Lauf gegen eine von Hand gefuehrte Tabelle.

    Dort gibt es noch keine Referenzspalte; die Zuordnung haengt an dem
    Link, den man beim Anlegen der Zeile eingefuegt hat.
    """
    ziel = tmp_path / "tracker.xlsx"
    mappe = Workbook()
    blatt = mappe.active
    kopf = {name: nr for nr, name in enumerate(felder.SPALTEN, start=1)}
    for name, nr in kopf.items():
        blatt.cell(row=1, column=nr, value=name)
    blatt.cell(row=2, column=kopf["Firma"], value="Beispiel GmbH")
    blatt.cell(row=2, column=kopf["Link zur Ausschreibung"], value=f"{felder.STELLENLINK}10001-1-S")
    blatt.cell(row=2, column=kopf["Kontakt"], value="Frau Meyer, 040 123")
    mappe.save(ziel)

    bericht = excel.schreibe(ziel, [zeile()])

    assert (bericht.neu, bericht.aktualisiert) == (0, 1)
    blatt, spalten = lies(ziel)
    assert blatt.max_row == 2
    assert blatt.cell(row=2, column=spalten["Kontakt"]).value == "Frau Meyer, 040 123"
    assert blatt.cell(row=2, column=spalten[felder.SPALTE_REFERENZ]).value == "10001-1-S"


def test_zuordnung_ueber_firma_und_position_ohne_link(tmp_path):
    ziel = tmp_path / "tracker.xlsx"
    mappe = Workbook()
    blatt = mappe.active
    kopf = {name: nr for nr, name in enumerate(felder.SPALTEN, start=1)}
    for name, nr in kopf.items():
        blatt.cell(row=1, column=nr, value=name)
    blatt.cell(row=2, column=kopf["Firma"], value="beispiel gmbh")
    blatt.cell(row=2, column=kopf["Position"], value="Data Engineer")
    mappe.save(ziel)

    bericht = excel.schreibe(ziel, [zeile()])

    assert (bericht.neu, bericht.aktualisiert) == (0, 1)


def test_eigene_spalten_und_reihenfolge_bleiben_erhalten(tmp_path):
    ziel = tmp_path / "tracker.xlsx"
    mappe = Workbook()
    blatt = mappe.active
    # Eigene Spalte vorn, Tracker-Spalten in anderer Reihenfolge.
    for spalte, name in enumerate(["Bauchgefühl", "Position", "Firma"], start=1):
        blatt.cell(row=1, column=spalte, value=name)
    blatt.cell(row=2, column=1, value="gut")
    blatt.cell(row=2, column=2, value="Data Engineer")
    blatt.cell(row=2, column=3, value="Beispiel GmbH")
    mappe.save(ziel)

    excel.schreibe(ziel, [zeile()])

    blatt, spalten = lies(ziel)
    assert blatt.max_row == 2
    assert spalten["Bauchgefühl"] == 1
    assert blatt.cell(row=2, column=1).value == "gut"
    # Fehlende Tracker-Spalten kommen rechts dazu, ohne die vorhandenen
    # zu verschieben.
    assert spalten["Quelle"] > 3
    assert blatt.cell(row=2, column=spalten["Standort"]).value == "20095 Hamburg"


def test_ueberschrift_mit_leerzeichen_legt_keine_zweite_spalte_an(tmp_path):
    ziel = tmp_path / "tracker.xlsx"
    mappe = Workbook()
    mappe.active.cell(row=1, column=1, value="Firma ")
    mappe.save(ziel)

    excel.schreibe(ziel, [zeile()])

    _, spalten = lies(ziel)
    assert [name for name in spalten if name.strip() == "Firma"] == ["Firma "]


def test_neue_anzeigen_haengen_unten_an_und_werden_durchnummeriert(tmp_path):
    ziel = tmp_path / "tracker.xlsx"
    excel.schreibe(ziel, [zeile(referenz="10001-1-S")])

    bericht = excel.schreibe(ziel, [zeile(referenz="10001-2-S", firma="Zweite GmbH")])

    assert (bericht.neu, bericht.aktualisiert) == (1, 0)
    blatt, spalten = lies(ziel)
    assert blatt.cell(row=3, column=spalten["Firma"]).value == "Zweite GmbH"
    assert [blatt.cell(row=r, column=spalten["Nr."]).value for r in (2, 3)] == [1, 2]


def test_zeilen_werden_nach_passung_sortiert(tmp_path):
    ziel = tmp_path / "tracker.xlsx"

    def bewertet(referenz, stufe, punkte):
        return zeile(referenz=referenz, firma=f"F-{referenz}", **{"Passung": stufe, "Punkte": punkte})

    zeilen = [
        bewertet("10001-1-S", "C – Randbereich", 20),
        bewertet("10001-2-S", "A – Volltreffer", 80),
        bewertet("10001-3-S", "A – Volltreffer", 95),
        bewertet("10001-4-S", "D – zu wenig Angaben", ""),
    ]

    excel.schreibe(ziel, zeilen, spalten=felder.SPALTEN_MIT_PASSUNG)

    blatt, spalten = lies(ziel)
    reihenfolge = [
        blatt.cell(row=r, column=spalten[felder.SPALTE_REFERENZ]).value for r in range(2, 6)
    ]
    assert reihenfolge == ["10001-3-S", "10001-2-S", "10001-1-S", "10001-4-S"]
    assert [blatt.cell(row=r, column=spalten["Nr."]).value for r in range(2, 6)] == [1, 2, 3, 4]


def test_entfernen_raeumt_nur_unbearbeitete_zeilen(tmp_path):
    ziel = tmp_path / "tracker.xlsx"
    excel.schreibe(
        ziel,
        [
            zeile(referenz="10001-1-S", firma="Senior AG"),
            zeile(referenz="10001-2-S", firma="Beworben GmbH", **{"Status": "Abgeschickt"}),
        ],
    )

    bericht = excel.schreibe(
        ziel,
        [zeile(referenz="10001-2-S", firma="Beworben GmbH", **{"Status": "Abgeschickt"})],
        entfernen={"10001-1-S", "10001-2-S"},
    )

    assert bericht.entfernt == 1
    blatt, spalten = lies(ziel)
    refs = [
        blatt.cell(row=r, column=spalten[felder.SPALTE_REFERENZ]).value
        for r in range(2, blatt.max_row + 1)
    ]
    assert refs == ["10001-2-S"]
    assert blatt.cell(row=2, column=spalten["Nr."]).value == 1


def test_ohne_passung_bleibt_die_fundreihenfolge(tmp_path):
    ziel = tmp_path / "tracker.xlsx"
    excel.schreibe(ziel, [zeile(referenz="10001-9-S", firma="Zzz GmbH")])

    excel.schreibe(ziel, [zeile(referenz="10001-1-S", firma="Aaa GmbH")])

    blatt, spalten = lies(ziel)
    reihenfolge = [
        blatt.cell(row=r, column=spalten[felder.SPALTE_REFERENZ]).value for r in (2, 3)
    ]
    assert reihenfolge == ["10001-9-S", "10001-1-S"]


def test_sicherung_wird_vor_dem_ueberschreiben_angelegt(tmp_path):
    ziel = tmp_path / "tracker.xlsx"
    excel.schreibe(ziel, [zeile()])

    bericht = excel.schreibe(ziel, [zeile()])

    assert bericht.sicherung == str(tmp_path / "tracker.sicherung.xlsx")
    assert (tmp_path / "tracker.sicherung.xlsx").exists()


def test_kein_unfertiges_zwischenergebnis_bleibt_liegen(tmp_path):
    ziel = tmp_path / "tracker.xlsx"

    excel.schreibe(ziel, [zeile()])

    assert not (tmp_path / "tracker.unfertig.xlsx").exists()


def test_referenzspalte_ist_ausgeblendet(tmp_path):
    ziel = tmp_path / "tracker.xlsx"

    excel.schreibe(ziel, [zeile()])

    blatt = load_workbook(ziel).active
    spalte = next(
        z.column_letter for z in blatt[1] if z.value == felder.SPALTE_REFERENZ
    )
    assert blatt.column_dimensions[spalte].hidden


def test_blattname_waehlt_das_richtige_arbeitsblatt(tmp_path):
    ziel = tmp_path / "tracker.xlsx"
    mappe = Workbook()
    mappe.active.title = "Notizen"
    mappe.create_sheet("Bewerbungen")
    mappe.save(ziel)

    excel.schreibe(ziel, [zeile()], blattname="Bewerbungen")

    mappe = load_workbook(ziel)
    assert mappe["Notizen"].max_row == 1
    blatt = mappe["Bewerbungen"]
    kopf = {z.value: z.column for z in blatt[1]}
    assert blatt.cell(row=2, column=kopf["Firma"]).value == "Beispiel GmbH"


def test_unbekannter_blattname_faellt_auf(tmp_path):
    ziel = tmp_path / "tracker.xlsx"
    Workbook().save(ziel)

    with pytest.raises(KeyError):
        excel.schreibe(ziel, [zeile()], blattname="Gibtsnicht")


# --- Status als Eingabefeld -------------------------------------------


def test_statusspalte_bekommt_ein_auswahlfeld(tmp_path):
    ziel = tmp_path / "tracker.xlsx"

    excel.schreibe(ziel, [zeile()])

    blatt = load_workbook(ziel).active
    pruefungen = list(blatt.data_validations.dataValidation)
    assert len(pruefungen) == 1
    for wert in status.AUSWAHL:
        assert wert in pruefungen[0].formula1


def test_auswahl_wird_zurueckgelesen(tmp_path):
    ziel = tmp_path / "tracker.xlsx"
    excel.schreibe(ziel, [zeile(referenz="10001-1-S")])

    mappe = load_workbook(ziel)
    _, spalten = lies(ziel)
    mappe.active.cell(row=2, column=spalten["Status"], value="Abgeschickt")
    mappe.save(ziel)

    assert excel.lies_status(ziel) == {"10001-1-S": "Abgeschickt"}


def test_lies_status_ohne_datei_ist_leer(tmp_path):
    assert excel.lies_status(tmp_path / "gibtsnicht.xlsx") == {}


def test_nicht_interessant_fliegt_raus_auch_wenn_bearbeitet(tmp_path):
    """Dass jemand die Zeile angefasst hat, ist hier gerade der Grund."""
    ziel = tmp_path / "tracker.xlsx"
    excel.schreibe(
        ziel,
        [
            zeile(referenz="10001-1-S", firma="Uninteressant AG",
                  **{"Status": "Nicht interessant"}),
            zeile(referenz="10001-2-S", firma="Bleibt GmbH"),
        ],
    )

    bericht = excel.schreibe(
        ziel,
        [zeile(referenz="10001-2-S", firma="Bleibt GmbH")],
        nicht_interessant={"10001-1-S"},
    )

    assert bericht.entfernt == 1
    blatt, spalten = lies(ziel)
    verbleibend = [
        blatt.cell(row=r, column=spalten[felder.SPALTE_REFERENZ]).value
        for r in range(2, blatt.max_row + 1)
    ]
    assert verbleibend == ["10001-2-S"]


def test_laufende_bewerbungen_stehen_oben_absagen_unten(tmp_path):
    ziel = tmp_path / "tracker.xlsx"

    def z(referenz, stufe, punkte, zustand=""):
        return zeile(
            referenz=referenz,
            firma=f"F-{referenz}",
            **{"Passung": stufe, "Punkte": punkte, "Status": zustand},
        )

    zeilen = [
        z("neu-a", "A – Volltreffer", 80),
        z("absage", "A – Volltreffer", 95, "Absage"),
        z("neu-b", "B – Naheliegend", 40),
        z("beworben", "C – Randbereich", 20, "Abgeschickt"),
        z("interview", "D – zu wenig Angaben", "", "Interview"),
    ]

    excel.schreibe(ziel, zeilen, spalten=felder.SPALTEN_MIT_PASSUNG)

    blatt, spalten = lies(ziel)
    reihenfolge = [
        blatt.cell(row=r, column=spalten[felder.SPALTE_REFERENZ]).value
        for r in range(2, 7)
    ]
    # Laufendes zuerst (weiteste Stufe oben), dann Neues nach Passung,
    # Erledigtes ganz unten - unabhaengig von der Punktzahl.
    assert reihenfolge == ["interview", "beworben", "neu-a", "neu-b", "absage"]
