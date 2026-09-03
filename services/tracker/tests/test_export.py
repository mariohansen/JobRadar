"""Tests der Datenbeschaffung.

Archiv und Schnittstelle werden durch Doppel ersetzt; geprueft wird vor
allem, dass die Detailansicht je Anzeige genau einmal abgerufen wird.
"""
from datetime import datetime, timezone

import pytest

from gemeinsam import jobdetail

from tracker import benefits, export as ex
from tracker.store import Eintrag

ERFASST = int(datetime(2025, 9, 1, 12, 0, tzinfo=timezone.utc).timestamp())

EINTRAG = Eintrag(
    referenznummer="10001-1-S",
    titel="Data Engineer",
    status="GEFUNDEN",
    erfasst_am=ERFASST,
    geaendert_am=None,
)


class FakeArchiv:
    def __init__(self, rohdaten=None, details=None):
        self._rohdaten = rohdaten or {}
        self._details = details or {}
        self.gemerkt = {}

    def rohdaten(self, referenznummer, erfasst_am):
        return self._rohdaten.get(referenznummer)

    def detail(self, referenznummer):
        return self._details.get(referenznummer)

    def merke_detail(self, referenznummer, inhalt):
        self.gemerkt[referenznummer] = inhalt
        self._details[referenznummer] = inhalt


def test_detail_kommt_aus_dem_zwischenspeicher_ohne_abruf():
    archiv = FakeArchiv(details={"10001-1-S": {"verguetung": "gut"}})

    def nicht_aufrufen(referenznummer):
        raise AssertionError("Die Schnittstelle haette nicht befragt werden duerfen")

    quellen = ex.Quellen(archiv, abruf=nicht_aufrufen)

    assert quellen.detail(EINTRAG) == {"verguetung": "gut"}


def test_frisches_detail_wird_gemerkt(monkeypatch):
    monkeypatch.setattr(ex.time, "sleep", lambda _: None)
    archiv = FakeArchiv()
    aufrufe = []

    def abruf(referenznummer):
        aufrufe.append(referenznummer)
        return {"stellenbeschreibung": "Text"}

    quellen = ex.Quellen(archiv, abruf=abruf)
    quellen.detail(EINTRAG)
    quellen.detail(EINTRAG)

    assert aufrufe == ["10001-1-S"]
    assert archiv.gemerkt == {"10001-1-S": {"stellenbeschreibung": "Text"}}


def test_zurueckgezogene_anzeige_wird_nicht_erneut_gefragt(monkeypatch):
    """Ein 404 ist der Normalfall und darf keine Dauerlast erzeugen."""
    monkeypatch.setattr(ex.time, "sleep", lambda _: None)
    archiv = FakeArchiv()
    aufrufe = []

    def abruf(referenznummer):
        aufrufe.append(referenznummer)
        return None

    quellen = ex.Quellen(archiv, abruf=abruf)
    assert quellen.detail(EINTRAG) == {}
    quellen.detail(EINTRAG)

    assert len(aufrufe) == 1


def test_unerreichbare_anzeige_bricht_den_export_nicht_ab():
    def abruf(referenznummer):
        raise jobdetail.JobdetailError("HTTP 500")

    quellen = ex.Quellen(FakeArchiv(), abruf=abruf)

    assert quellen.detail(EINTRAG) == {}


def test_ohne_details_wird_gar_nicht_abgerufen():
    def nicht_aufrufen(referenznummer):
        raise AssertionError("Es sollte kein Abruf stattfinden")

    quellen = ex.Quellen(FakeArchiv(), mit_details=False, abruf=nicht_aufrufen)

    assert quellen.detail(EINTRAG) == {}


def test_details_erneuern_uebergeht_den_zwischenspeicher(monkeypatch):
    monkeypatch.setattr(ex.time, "sleep", lambda _: None)
    archiv = FakeArchiv(details={"10001-1-S": {"verguetung": "alt"}})

    quellen = ex.Quellen(archiv, details_erneuern=True, abruf=lambda r: {"verguetung": "neu"})

    assert quellen.detail(EINTRAG) == {"verguetung": "neu"}


def test_baue_zeilen_meldet_den_fortschritt():
    archiv = FakeArchiv(rohdaten={"10001-1-S": {"firma": "Beispiel GmbH"}})
    quellen = ex.Quellen(archiv, mit_details=False)
    stand = []

    zeilen = ex.baue_zeilen([EINTRAG], quellen, fortschritt=lambda i, n: stand.append((i, n)))

    assert stand == [(1, 1)]
    assert zeilen[0]["Firma"] == "Beispiel GmbH"


def test_baue_zeilen_bewertet_mit_profil():
    from gemeinsam import profil as pr

    profil = pr.Profil(faehigkeiten={"Python": 3, "SQL": 2})
    archiv = FakeArchiv(
        details={"10001-1-S": {"stellenangebotsBeschreibung": "Python, SQL, Airflow, ETL"}}
    )
    quellen = ex.Quellen(archiv, abruf=lambda r: {})

    zeilen = ex.baue_zeilen([EINTRAG], quellen, profil=profil)

    assert zeilen[0]["Passung"].startswith(("A", "B", "C"))
    assert "Python" in zeilen[0]["Treffer"]


# --- Benefits ---------------------------------------------------------


def test_benefits_erkennt_die_gaengigen_begriffe():
    text = (
        "Wir bieten ein Jobticket, betriebliche Altersvorsorge, "
        "30 Tage Urlaub und flexible Arbeitszeiten."
    )

    gefunden = benefits.finde(text)

    assert "Jobticket / ÖPNV" in gefunden
    assert "Betriebliche Altersvorsorge" in gefunden
    assert "30+ Tage Urlaub" in gefunden
    assert "Gleitzeit" in gefunden


def test_benefits_ignoriert_gross_und_kleinschreibung_und_umlautersatz():
    assert "Jobrad / Bikeleasing" in benefits.finde("JOBRAD moeglich")
    assert "Vermögenswirksame Leistungen" in benefits.finde("vermoegenswirksame leistungen")


def test_urlaub_erst_ab_dreissig_tagen():
    """Darunter ist es der gesetzliche Rahmen und kein Argument."""
    assert benefits.finde("25 Urlaubstage") == []
    assert benefits.finde("30 Urlaubstage") == ["30+ Tage Urlaub"]


def test_leerer_text_ergibt_nichts():
    assert benefits.finde("") == []
    assert benefits.als_text("Ein Text ohne Zusatzleistungen") == ""


# --- Verdrahtung des Befehls -----------------------------------------


def test_befehl_export_schreibt_die_datei(tmp_path, monkeypatch):
    """Ein Lauf ueber main, ohne AWS und ohne Netz."""
    from openpyxl import load_workbook

    from tracker import main as m

    class FakeStore:
        def __init__(self, tabelle):
            pass

        def liste(self, nur_status=None):
            return iter([EINTRAG])

    monkeypatch.setenv("DYNAMODB_TABLE_SEEN_JOBS", "tabelle")
    monkeypatch.setenv("S3_BUCKET_RAW_ARCHIVE", "eimer")
    monkeypatch.setattr(m, "Store", FakeStore)
    monkeypatch.setattr(
        "gemeinsam.archiv.Archiv",
        lambda bucket: FakeArchiv(rohdaten={"10001-1-S": {"firma": "Beispiel GmbH"}}),
    )

    ziel = tmp_path / "tracker.xlsx"
    assert m.main(["export", "--datei", str(ziel), "--ohne-details"]) == 0

    blatt = load_workbook(ziel).active
    kopf = {z.value: z.column for z in blatt[1]}
    assert blatt.cell(row=2, column=kopf["Firma"]).value == "Beispiel GmbH"


def test_fehlender_bucket_nennt_den_terraform_befehl(monkeypatch):
    from tracker import main as m

    monkeypatch.delenv("S3_BUCKET_RAW_ARCHIVE", raising=False)

    with pytest.raises(SystemExit) as fehler:
        m._bucketname()

    assert "archive_bucket_name" in str(fehler.value)


# --- Titel-Ausschluss ------------------------------------------------


def _eintrag(referenz, titel, status="GEFUNDEN"):
    return Eintrag(referenz, titel, status, ERFASST, None)


def test_teile_aussortierte_trennt_senior_und_lead(monkeypatch):
    from tracker import main as m

    monkeypatch.delenv("MATCH_AUSSCHLUSS", raising=False)
    eintraege = [
        _eintrag("1-S", "Data Engineer (m/w/d)"),
        _eintrag("2-S", "Senior Data Engineer"),
        _eintrag("3-S", "Team Lead Data Platform"),
    ]

    behalten, aussortiert = m._teile_aussortierte(eintraege, mit_aussortierten=False)

    assert [e.referenznummer for e in behalten] == ["1-S"]
    assert aussortiert == {"senior": ["2-S"], "lead": ["3-S"]}


def test_teile_aussortierte_verschont_laufende_bewerbungen(monkeypatch):
    from tracker import main as m

    monkeypatch.delenv("MATCH_AUSSCHLUSS", raising=False)
    eintraege = [_eintrag("2-S", "Senior Data Engineer", status="BEWORBEN")]

    behalten, aussortiert = m._teile_aussortierte(eintraege, mit_aussortierten=False)

    assert [e.referenznummer for e in behalten] == ["2-S"]
    assert aussortiert == {}


def test_teile_aussortierte_kann_abgeschaltet_werden(monkeypatch):
    from tracker import main as m

    eintraege = [_eintrag("2-S", "Senior Data Engineer")]

    behalten, aussortiert = m._teile_aussortierte(eintraege, mit_aussortierten=True)

    assert [e.referenznummer for e in behalten] == ["2-S"]
    assert aussortiert == {}


# --- Status aus der Tabelle zurueck --------------------------------


class FakeStore:
    """Merkt sich Statusaenderungen, ohne DynamoDB anzufassen."""

    def __init__(self, eintraege):
        self.eintraege = list(eintraege)
        self.gesetzt = []

    def setze_status(self, referenznummer, neuer_status):
        self.gesetzt.append((referenznummer, neuer_status))


def _tabellenstatus(monkeypatch, inhalt):
    from tracker import main as m

    monkeypatch.setattr("tracker.excel.lies_status", lambda datei, blatt=None: inhalt)
    return m


def test_auswahl_wandert_nach_dynamodb(monkeypatch):
    m = _tabellenstatus(monkeypatch, {"10001-1-S": "Abgeschickt"})
    store = FakeStore([])
    eintraege = [_eintrag("10001-1-S", "Data Engineer")]

    aktualisiert, geaendert, unbekannt = m._uebernimm_status(
        store, eintraege, "egal.xlsx", None
    )

    assert store.gesetzt == [("10001-1-S", "BEWORBEN")]
    assert geaendert == 1
    assert unbekannt == []
    # Der Eintrag traegt den neuen Status auch im Speicher.
    assert aktualisiert[0].status == "BEWORBEN"


def test_unveraenderter_status_wird_nicht_geschrieben(monkeypatch):
    m = _tabellenstatus(monkeypatch, {"10001-1-S": "Abgeschickt"})
    store = FakeStore([])
    eintraege = [_eintrag("10001-1-S", "Data Engineer", status="BEWORBEN")]

    _, geaendert, _ = m._uebernimm_status(store, eintraege, "egal.xlsx", None)

    assert store.gesetzt == []
    assert geaendert == 0


def test_leere_zelle_bedeutet_gefunden(monkeypatch):
    m = _tabellenstatus(monkeypatch, {"10001-1-S": None})
    store = FakeStore([])
    eintraege = [_eintrag("10001-1-S", "Data Engineer")]

    _, geaendert, _ = m._uebernimm_status(store, eintraege, "egal.xlsx", None)

    assert store.gesetzt == []
    assert geaendert == 0


def test_unbekannter_zellinhalt_wird_gemeldet_nicht_geraten(monkeypatch):
    m = _tabellenstatus(monkeypatch, {"10001-1-S": "Vielleicht"})
    store = FakeStore([])
    eintraege = [_eintrag("10001-1-S", "Data Engineer")]

    aktualisiert, geaendert, unbekannt = m._uebernimm_status(
        store, eintraege, "egal.xlsx", None
    )

    assert store.gesetzt == []
    assert unbekannt == ["Vielleicht"]
    assert aktualisiert[0].status == "GEFUNDEN"


def test_ohne_tabelle_passiert_nichts(monkeypatch):
    m = _tabellenstatus(monkeypatch, {})
    store = FakeStore([])
    eintraege = [_eintrag("10001-1-S", "Data Engineer")]

    aktualisiert, geaendert, unbekannt = m._uebernimm_status(
        store, eintraege, "gibtsnicht.xlsx", None
    )

    assert (aktualisiert, geaendert, unbekannt) == (eintraege, 0, [])
