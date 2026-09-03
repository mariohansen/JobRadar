"""Schreiben und Fortschreiben der Tracker-Tabelle.

Der Export ergaenzt eine vorhandene Datei, statt sie zu ersetzen. Das ist
die eigentliche Anforderung: die Spalten mit den Anzeigendaten kommen aus
der Pipeline, Abgabedatum, Fristen und Notizen stehen daneben und sind
Handarbeit. Ein Export, der die Datei jedes Mal neu schreibt, waere nach
der ersten eigenen Eintragung unbrauchbar.

Drei Stufen, siehe felder.AUTOMATISCH und felder.VORSCHLAG:

* automatisch - wird bei jedem Lauf ueberschrieben,
* Vorschlag   - wird nur eingetragen, solange die Zelle leer ist,
* von Hand    - wird nie angefasst.

Zugeordnet wird ueber die Ueberschriften der ersten Zeile, nicht ueber
die Spaltenposition. Eine umsortierte oder um eigene Spalten erweiterte
Tabelle bleibt damit lesbar, und fremde Spalten bleiben unberuehrt.
"""
from __future__ import annotations

import logging
import math
import os
import shutil
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Iterable

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import FormulaRule
from openpyxl.worksheet.datavalidation import DataValidation

from . import status as st

from . import felder

log = logging.getLogger(__name__)

BLATTNAME = "Bewerbungen"

# Nur fuer neu angelegte Dateien. Eine bestehende Tabelle behaelt ihre
# eigene Gestaltung - der Export ist dort Datenlieferant, nicht Setzer.
# Bis zu welcher Zeile das Auswahlfeld reicht. Grosszuegig, damit es
# auch fuer Zeilen gilt, die erst spaeter dazukommen.
LETZTE_ZEILE = 5000

BREITEN = {
    "Nr.": 5,
    "Passung": 19,
    "Punkte": 7,
    "Status": 16,
    "Firma": 30,
    "Position": 40,
    "Link zur Ausschreibung": 46,
    "Standort": 22,
    "Entfernung (km)": 9,
    "Benefits": 46,
    "Homeoffice-Modell": 21,
    "Gehalt": 18,
    "Alter (Tage)": 8,
    "Treffer": 42,
    "Lücken": 42,
    "Kontakt": 28,
    "Quelle": 13,
}

# Spalten, die Zahlen fuehren: mittig, ohne Umbruch.
ZAHLENSPALTEN = ("Nr.", "Punkte", "Entfernung (km)", "Alter (Tage)")

ZAHLENFORMATE = {
    "Nr.": "0",
    "Punkte": "0",
    "Alter (Tage)": "0",
    "Entfernung (km)": "0.0",
}

# Punkt je Textzeile bei der vorgegebenen Schriftgroesse.
ZEILENHOEHE = 15

# Notbremse: eine einzelne ausufernde Zelle soll nicht die ganze Tabelle
# auf Handtuchhoehe ziehen.
HOECHSTE_ZEILEN = 8

KOPF_FARBE = PatternFill("solid", fgColor="1F3864")
KOPF_SCHRIFT = Font(bold=True, color="FFFFFF")
LINK_SCHRIFT = Font(color="0563C1", underline="single")

# Gedeckte Toene: die Tabelle soll gegliedert aussehen, nicht bunt.
PASSUNGSFARBEN = {
    "A": PatternFill("solid", fgColor="C6EFCE"),
    "B": PatternFill("solid", fgColor="E2EFDA"),
    "C": PatternFill("solid", fgColor="F2F2F2"),
    "D": PatternFill("solid", fgColor="FBFBFB"),
}

STATUSFARBEN = {
    "Zusage": PatternFill("solid", fgColor="A9D08E"),
    "Interview": PatternFill("solid", fgColor="BDD7EE"),
    "Abgeschickt": PatternFill("solid", fgColor="FFE699"),
    "Absage": PatternFill("solid", fgColor="E7E6E6"),
}


@dataclass(frozen=True)
class Bericht:
    datei: str
    neu: int
    aktualisiert: int
    sicherung: str | None
    entfernt: int = 0


def _schluessel(ueberschrift: Any) -> str:
    """Vergleichsform einer Ueberschrift.

    Von Hand gepflegte Tabellen enthalten regelmaessig ein nachgestelltes
    Leerzeichen oder einen Zeilenumbruch in der Kopfzeile. Ohne
    Normalisierung wuerde die Spalte dann ein zweites Mal angelegt.
    """
    if not isinstance(ueberschrift, str):
        return ""
    return " ".join(ueberschrift.split()).casefold()


def _referenz_aus_link(wert: Any) -> str:
    if not isinstance(wert, str) or felder.STELLENLINK not in wert:
        return ""
    return wert.rsplit("/", 1)[-1].strip()


def _paarschluessel(firma: Any, position: Any) -> str:
    teile = [" ".join(str(t).split()).casefold() for t in (firma, position) if t]
    return " | ".join(teile) if len(teile) == 2 else ""


class Tabelle:
    """Ein Arbeitsblatt mit den Tracker-Spalten."""

    def __init__(self, blatt) -> None:
        self._blatt = blatt
        self._spalten: dict[str, int] = {}
        self._lies_kopfzeile()

    def _lies_kopfzeile(self) -> None:
        for nummer, zelle in enumerate(self._blatt[1], start=1):
            name = _schluessel(zelle.value)
            if name and name not in self._spalten:
                self._spalten[name] = nummer

    def ergaenze_spalten(self, spalten: Iterable[str]) -> None:
        """Legt fehlende Spalten rechts an, ohne bestehende zu verschieben."""
        naechste = max(self._spalten.values(), default=0) + 1
        for name in (*spalten, felder.SPALTE_REFERENZ):
            if _schluessel(name) in self._spalten:
                continue
            self._blatt.cell(row=1, column=naechste, value=name)
            self._spalten[_schluessel(name)] = naechste
            naechste += 1

    def spalte(self, name: str) -> int:
        return self._spalten[_schluessel(name)]

    def datenzeilen(self) -> list[int]:
        """Zeilennummern mit Inhalt, ohne die Kopfzeile."""
        return [
            zeile[0].row
            for zeile in self._blatt.iter_rows(min_row=2)
            if any(zelle.value not in (None, "") for zelle in zeile)
        ]

    def finde(self, referenznummer: str, firma: str, position: str) -> int | None:
        """Zeile zu einer Anzeige, ueber drei Wege.

        Beim ersten Lauf gegen eine von Hand gefuehrte Tabelle gibt es
        noch keine Referenzspalte. Dann traegt der Link die Nummer, und
        wenn auch der fehlt, bleibt der Vergleich von Firma und Position.
        """
        spalte_referenz = self.spalte(felder.SPALTE_REFERENZ)
        spalte_link = self.spalte("Link zur Ausschreibung")
        spalte_firma = self.spalte("Firma")
        spalte_position = self.spalte("Position")
        gesucht = _paarschluessel(firma, position)

        ueber_paar: int | None = None
        for nummer in self.datenzeilen():
            wert = self._blatt.cell(row=nummer, column=spalte_referenz).value
            if isinstance(wert, str) and wert.strip() == referenznummer:
                return nummer
            if _referenz_aus_link(self._blatt.cell(row=nummer, column=spalte_link).value) == referenznummer:
                return nummer
            if gesucht and ueber_paar is None:
                vorhanden = _paarschluessel(
                    self._blatt.cell(row=nummer, column=spalte_firma).value,
                    self._blatt.cell(row=nummer, column=spalte_position).value,
                )
                if vorhanden == gesucht:
                    ueber_paar = nummer

        return ueber_paar

    def naechste_zeile(self) -> int:
        return max(self.datenzeilen(), default=1) + 1

    def schreibe_zeile(self, nummer: int, werte: dict[str, str], ueberschreiben: bool) -> None:
        for name, wert in werte.items():
            if _schluessel(name) not in self._spalten:
                # Eine Spalte, die diese Tabelle nicht fuehrt - etwa aus
                # einem aelteren Schema. Kein Grund, den Export abzubrechen.
                continue
            zelle = self._blatt.cell(row=nummer, column=self.spalte(name))

            if name in felder.AUTOMATISCH or name == felder.SPALTE_REFERENZ:
                pass
            elif name in felder.VORSCHLAG:
                # Eine eigene Eintragung wiegt schwerer als ein Vorschlag.
                if zelle.value not in (None, "") and not ueberschreiben:
                    continue
            else:
                continue

            if wert in ("", None) and zelle.value not in (None, "") and name != "Status":
                # Ein leerer Wert entsteht auch, wenn eine Quelle diesmal
                # fehlt. Vorhandenes dafuer zu loeschen waere Datenverlust.
                # Der Status ist die Ausnahme: er kommt aus DynamoDB und
                # ist dort auch dann richtig, wenn er leer ist.
                continue

            zelle.value = wert
            if name == "Link zur Ausschreibung" and wert:
                zelle.hyperlink = wert
                zelle.font = LINK_SCHRIFT

    def lies_status(self) -> dict[str, Any]:
        """Referenznummer -> Zellinhalt der Statusspalte.

        Die Spalte ist das Eingabefeld des Benutzers; der Export liest
        sie zurueck, bevor er irgendetwas schreibt.
        """
        if _schluessel("Status") not in self._spalten:
            return {}

        sp_ref = self.spalte(felder.SPALTE_REFERENZ)
        sp_status = self.spalte("Status")
        gelesen: dict[str, Any] = {}
        for nummer in self.datenzeilen():
            referenz = self._blatt.cell(row=nummer, column=sp_ref).value
            if isinstance(referenz, str) and referenz.strip():
                gelesen[referenz.strip()] = self._blatt.cell(row=nummer, column=sp_status).value
        return gelesen

    def entferne(self, nur_unbearbeitet: set[str], immer: set[str] = frozenset()) -> int:
        """Loescht Datenzeilen zu diesen Referenzen.

        `nur_unbearbeitet` trifft Anzeigen, die jetzt unter den
        Titel-Ausschluss fallen - dort bleibt stehen, woran schon Arbeit
        haengt. `immer` trifft die als "Nicht interessant" markierten:
        die sollen weg, und zwar genau deshalb, weil jemand sie
        angefasst hat.

        Statt einzelner delete_rows - die in openpyxl Formatreste
        hinterlassen - werden die verbleibenden Zeilen lueckenlos neu
        geschrieben und der Rest entfernt.
        """
        if not nur_unbearbeitet and not immer:
            return 0

        zeilen = self.datenzeilen()
        if not zeilen:
            return 0

        sp_ref = self.spalte(felder.SPALTE_REFERENZ)
        sp_status = self._spalten.get(_schluessel("Status"))
        sp_link = self._spalten.get(_schluessel("Link zur Ausschreibung"))
        breite = self._blatt.max_column

        behalten: list[list[Any]] = []
        for nummer in zeilen:
            wert = self._blatt.cell(row=nummer, column=sp_ref).value
            referenz = wert.strip() if isinstance(wert, str) else ""
            if referenz and referenz in immer:
                continue
            if referenz and referenz in nur_unbearbeitet:
                zelle = self._blatt.cell(row=nummer, column=sp_status) if sp_status else None
                inhalt = zelle.value if zelle is not None else None
                if inhalt in (None, "") or st.aus_tabelle(inhalt) == st.GEFUNDEN:
                    continue
            behalten.append(
                [self._blatt.cell(row=nummer, column=spalte).value for spalte in range(1, breite + 1)]
            )

        entfernt = len(zeilen) - len(behalten)
        if not entfernt:
            return 0

        erste = zeilen[0]
        for versatz, werte in enumerate(behalten):
            for spalte, wert in enumerate(werte, start=1):
                zelle = self._blatt.cell(row=erste + versatz, column=spalte, value=wert)
                if spalte == sp_link and wert:
                    zelle.hyperlink = wert
                    zelle.font = LINK_SCHRIFT

        # Die freigewordenen Zeilen am Ende ganz entfernen. Hier steht
        # keine Anzeige mehr, die verschoben werden koennte.
        naechste_freie = erste + len(behalten)
        ueberzaehlig = self._blatt.max_row - naechste_freie + 1
        if ueberzaehlig > 0:
            self._blatt.delete_rows(naechste_freie, ueberzaehlig)

        return entfernt

    def sortiere(self) -> None:
        """Datenzeilen ordnen: laufende Bewerbungen, Neues, Erledigtes.

        Drei Gruppen, damit die Tabelle von oben nach unten das
        wiedergibt, worum man sich kuemmern muss:

        1. laufende Bewerbungen - zuerst die weiteste (Zusage vor
           Interview vor Abgeschickt). Sie zwischen zweihundert
           unbearbeiteten Anzeigen zu suchen waere muehsam.
        2. alles Unberuehrte, nach Passung - beste Uebereinstimmung oben.
        3. Absagen. Erledigt, aber nicht vergessen.

        Sortiert wird die volle Zeilenbreite, damit fremde Spalten bei
        ihrer Anzeige bleiben. Zeilenbezogene Handformatierung und
        Formeln mit Zeilenbezug ueberstehen das nicht - dafuer liegt die
        Sicherung daneben.
        """
        zeilen = self.datenzeilen()
        if len(zeilen) < 2:
            return

        sp_passung = self._spalten.get(_schluessel("Passung"))
        sp_status = self._spalten.get(_schluessel("Status"))
        if sp_passung is None and sp_status is None:
            # Kein Kriterium - dann bleibt die Fundreihenfolge.
            return

        breite = self._blatt.max_column
        sp_punkte = self._spalten.get(_schluessel("Punkte"))
        sp_alter = self._spalten.get(_schluessel("Alter (Tage)"))
        sp_firma = self._spalten.get(_schluessel("Firma"))
        sp_link = self._spalten.get(_schluessel("Link zur Ausschreibung"))

        daten = [
            [self._blatt.cell(row=nummer, column=spalte).value for spalte in range(1, breite + 1)]
            for nummer in zeilen
        ]

        def gruppe(werte: list[Any]) -> tuple[int, int]:
            """Welcher Block, und innerhalb dessen welcher Rang."""
            if sp_status is None:
                return (1, 0)
            zustand = st.aus_tabelle(werte[sp_status - 1]) or st.GEFUNDEN
            if zustand in st.LAEUFT:
                # Weiteste zuerst: Zusage vor Interview vor Abgeschickt.
                return (0, -st.ALLE.index(zustand))
            if zustand == st.ABSAGE:
                return (2, 0)
            return (1, 0)

        def schluessel(werte: list[Any]) -> tuple:
            block, rang = gruppe(werte)
            if sp_passung is None:
                # Ohne Bewertung gibt es innerhalb einer Gruppe kein
                # Kriterium. Die Sortierung ist stabil, damit bleibt dort
                # die Fundreihenfolge stehen.
                return (block, rang)

            # "A – …" < "B – …" < … < "D – …" ergibt schon die richtige
            # Reihenfolge der Stufen; leere Zellen ganz nach hinten.
            stufe = werte[sp_passung - 1] or "Z"
            punkte = werte[sp_punkte - 1] if sp_punkte else None
            punkte = punkte if isinstance(punkte, (int, float)) else -1
            alter = werte[sp_alter - 1] if sp_alter else None
            alter = alter if isinstance(alter, (int, float)) else 10**9
            firma = (werte[sp_firma - 1] if sp_firma else "") or ""
            return (block, rang, str(stufe), -punkte, alter, str(firma).casefold())

        daten.sort(key=schluessel)

        for nummer, werte in zip(zeilen, daten):
            for spalte, wert in enumerate(werte, start=1):
                zelle = self._blatt.cell(row=nummer, column=spalte, value=wert)
                if spalte == sp_link and wert:
                    zelle.hyperlink = wert
                    zelle.font = LINK_SCHRIFT

    def auswahlfeld(self) -> None:
        """Legt das Auswahlmenue auf die Statusspalte.

        Damit ist die Tabelle das Eingabefeld: anklicken statt tippen,
        und ein Tippfehler kommt gar nicht erst zustande. Die Auswahl
        liest der naechste Export zurueck.
        """
        if _schluessel("Status") not in self._spalten:
            return

        buchstabe = get_column_letter(self.spalte("Status"))
        pruefung = DataValidation(
            type="list",
            formula1='"' + ",".join(st.AUSWAHL) + '"',
            # Leer heisst: noch nichts entschieden.
            allow_blank=True,
            showErrorMessage=True,
        )
        pruefung.errorTitle = "Unbekannter Status"
        pruefung.error = "Bitte einen Wert aus der Liste waehlen."
        pruefung.promptTitle = "Status"
        pruefung.prompt = "Leer lassen, solange nichts entschieden ist."
        self._blatt.add_data_validation(pruefung)
        pruefung.add(f"{buchstabe}2:{buchstabe}{LETZTE_ZEILE}")

    def nummeriere(self) -> None:
        spalte = self.spalte("Nr.")
        for laufend, nummer in enumerate(self.datenzeilen(), start=1):
            self._blatt.cell(row=nummer, column=spalte, value=laufend)

    def verstecke_referenz(self) -> None:
        buchstabe = get_column_letter(self.spalte(felder.SPALTE_REFERENZ))
        self._blatt.column_dimensions[buchstabe].hidden = True

    def _hoechster_bedarf(self) -> int:
        """Wie viele Textzeilen die vollste Zelle der Tabelle braucht.

        Excel rechnet die Hoehe umbrochener Zellen nicht selbst aus, wenn
        die Hoehe gesetzt ist - und ohne gesetzte Hoehe sind die Zeilen
        unterschiedlich hoch. Deshalb wird der Bedarf hier geschaetzt:
        Zeichen geteilt durch Spaltenbreite, aufgerundet, ausdrueckliche
        Umbrueche mitgezaehlt.

        Die Schaetzung faellt bewusst grosszuegig aus - eine Zeile zu
        hoch faellt nicht auf, eine zu niedrig schneidet Text ab.
        """
        bedarf = 1
        for name, breite in BREITEN.items():
            schluessel = _schluessel(name)
            if schluessel not in self._spalten:
                continue
            spalte = self._spalten[schluessel]
            # Ein Zeichen Puffer fuer den Zellrand.
            passt = max(int(breite) - 1, 1)
            for nummer in self.datenzeilen():
                wert = self._blatt.cell(row=nummer, column=spalte).value
                if wert in (None, ""):
                    continue
                for absatz in str(wert).split("\n"):
                    bedarf = max(bedarf, math.ceil(len(absatz) / passt))
        return min(bedarf, HOECHSTE_ZEILEN)

    def gestalte(self) -> None:
        """Kopfzeile, Breiten, Hoehen, Farben und Filter.

        Laeuft bei **jedem** Export, nicht nur beim Anlegen: eine
        Tabelle, deren Zeilen nach jedem Lauf anders hoch sind, liest
        sich schlecht. Wer eigene Spalten hat, behaelt sie - nur die
        bekannten werden vermessen.
        """
        letzte = max(self.datenzeilen(), default=1)
        breiteste = max(self._spalten.values())

        for name, spalte in self._spalten.items():
            zelle = self._blatt.cell(row=1, column=spalte)
            zelle.fill = KOPF_FARBE
            zelle.font = KOPF_SCHRIFT
            zelle.alignment = Alignment(
                vertical="center", horizontal="center", wrap_text=True
            )

        for name, breite in BREITEN.items():
            if _schluessel(name) not in self._spalten:
                continue
            buchstabe = get_column_letter(self.spalte(name))
            self._blatt.column_dimensions[buchstabe].width = breite

        # Alle Datenzellen umbrechen und oben ausrichten, Zahlen mittig.
        for nummer in self.datenzeilen():
            for spalte in range(1, breiteste + 1):
                zelle = self._blatt.cell(row=nummer, column=spalte)
                mittig = spalte in self._zahlenspalten()
                zelle.alignment = Alignment(
                    vertical="top",
                    horizontal="center" if mittig else "left",
                    wrap_text=not mittig,
                )

        for name, format in ZAHLENFORMATE.items():
            if _schluessel(name) not in self._spalten:
                continue
            spalte = self.spalte(name)
            for nummer in self.datenzeilen():
                self._blatt.cell(row=nummer, column=spalte).number_format = format

        # Eine Hoehe fuer alle - so hoch wie die vollste Zelle es braucht.
        hoehe = self._hoechster_bedarf() * ZEILENHOEHE
        for nummer in self.datenzeilen():
            self._blatt.row_dimensions[nummer].height = hoehe

        self._blatt.row_dimensions[1].height = 32
        self._faerbe(letzte)

        # Erste Spalten stehen lassen: beim Blaettern nach rechts soll
        # sichtbar bleiben, um welche Anzeige es geht.
        self._blatt.freeze_panes = self._einfrierpunkt()

        # Ueber den ganzen Bereich, nicht nur die Kopfzeile - sonst
        # filtert Excel nichts, es zeigt nur die Pfeile.
        self._blatt.auto_filter.ref = f"A1:{get_column_letter(breiteste)}{letzte}"

    def _zahlenspalten(self) -> set[int]:
        return {
            self._spalten[_schluessel(name)]
            for name in ZAHLENSPALTEN
            if _schluessel(name) in self._spalten
        }

    def _einfrierpunkt(self) -> str:
        """Links von hier bleibt beim Blaettern stehen."""
        # Bis einschliesslich Position - Firma allein sagt zu wenig.
        for name in ("Position", "Status", "Nr."):
            if _schluessel(name) in self._spalten:
                return f"{get_column_letter(self.spalte(name) + 1)}2"
        return "A2"

    def _faerbe(self, letzte: int) -> None:
        """Bedingte Formatierung fuer Passung und Status.

        Als Regel und nicht als feste Farbe: waehlt jemand im
        Auswahlfeld einen anderen Status, faerbt Excel die Zeile sofort
        um, ohne dass ein Export dazwischen muss.
        """
        if letzte < 2:
            return

        if _schluessel("Passung") in self._spalten:
            buchstabe = get_column_letter(self.spalte("Passung"))
            bereich = f"{buchstabe}2:{buchstabe}{letzte}"
            for anfang, farbe in PASSUNGSFARBEN.items():
                self._blatt.conditional_formatting.add(
                    bereich,
                    FormulaRule(
                        formula=[f'LEFT(${buchstabe}2,1)="{anfang}"'], fill=farbe
                    ),
                )

        if _schluessel("Status") in self._spalten:
            buchstabe = get_column_letter(self.spalte("Status"))
            bereich = f"{buchstabe}2:{buchstabe}{letzte}"
            for wert, farbe in STATUSFARBEN.items():
                self._blatt.conditional_formatting.add(
                    bereich,
                    FormulaRule(formula=[f'${buchstabe}2="{wert}"'], fill=farbe),
                )


def _oeffne(pfad: Path, blattname: str | None, spalten: Iterable[str]) -> tuple[Any, Any, bool]:
    if pfad.exists():
        mappe = load_workbook(pfad)
        blatt = mappe[blattname] if blattname else mappe.active
        return mappe, blatt, False

    mappe = Workbook()
    blatt = mappe.active
    blatt.title = blattname or BLATTNAME
    for spalte, name in enumerate((*spalten, felder.SPALTE_REFERENZ), start=1):
        blatt.cell(row=1, column=spalte, value=name)
    return mappe, blatt, True


def lies_status(pfad: str | Path, blattname: str | None = None) -> dict[str, Any]:
    """Die Statusauswahl aus einer vorhandenen Tabelle.

    Wird vor dem Schreiben aufgerufen: was in der Tabelle steht, ist die
    Entscheidung des Benutzers und gehoert nach DynamoDB, bevor der
    Export von dort zurueckschreibt. Fehlt die Datei, gibt es nichts
    abzugleichen.
    """
    ziel = Path(pfad)
    if not ziel.exists():
        return {}

    mappe = load_workbook(ziel, read_only=False, data_only=True)
    try:
        blatt = mappe[blattname] if blattname else mappe.active
        return Tabelle(blatt).lies_status()
    finally:
        mappe.close()


def _sichere(pfad: Path) -> str | None:
    """Kopie der bestehenden Datei, bevor sie neu geschrieben wird.

    openpyxl liest die Mappe ein und schreibt sie vollstaendig neu.
    Diagramme, Bilder und Pivot-Tabellen ueberleben das nicht. Bei einer
    ueber Monate gefuehrten Bewerbungsuebersicht ist eine Sicherung die
    Muehe wert.
    """
    if not pfad.exists():
        return None
    ziel = pfad.with_name(f"{pfad.stem}.sicherung{pfad.suffix}")
    shutil.copy2(pfad, ziel)
    return str(ziel)


def schreibe(
    pfad: str | Path,
    zeilen: Iterable[dict[str, Any]],
    blattname: str | None = None,
    ueberschreiben: bool = False,
    spalten: Iterable[str] | None = None,
    entfernen: set[str] | None = None,
    nicht_interessant: set[str] | None = None,
) -> Bericht:
    """Schreibt die Zeilen in die Tabelle.

    `spalten` bestimmt, welche Ueberschriften angelegt werden. Ohne
    Faehigkeitsprofil sind das die Tracker-Spalten, mit Profil kommen
    die Passungsspalten dazu - eine leere Passungsspalte anzulegen, die
    nie gefuellt wird, waere nur Ballast.

    `entfernen` nennt Referenznummern, deren Zeile raus soll, wenn sie
    noch unbearbeitet ist - Anzeigen, die jetzt unter den Ausschluss
    fallen. Eine laufende Bewerbung bleibt unangetastet.
    `nicht_interessant` fliegt dagegen immer raus: dass jemand die
    Anzeige angefasst hat, ist hier gerade der Grund.
    """
    spalten = tuple(spalten) if spalten is not None else felder.SPALTEN
    ziel = Path(pfad)
    sicherung = _sichere(ziel)
    mappe, blatt, ist_neu = _oeffne(ziel, blattname, spalten)

    tabelle = Tabelle(blatt)
    tabelle.ergaenze_spalten(spalten)

    neu = 0
    aktualisiert = 0
    for werte in zeilen:
        nummer = tabelle.finde(
            werte[felder.SPALTE_REFERENZ], werte.get("Firma", ""), werte.get("Position", "")
        )
        if nummer is None:
            nummer = tabelle.naechste_zeile()
            neu += 1
        else:
            aktualisiert += 1
        tabelle.schreibe_zeile(nummer, werte, ueberschreiben)

    entfernt = tabelle.entferne(entfernen or set(), nicht_interessant or set())
    tabelle.sortiere()
    tabelle.nummeriere()
    tabelle.verstecke_referenz()
    tabelle.auswahlfeld()
    tabelle.gestalte()

    # Erst vollstaendig danebenschreiben, dann ersetzen. Bricht der Lauf
    # mittendrin ab, steht die alte Datei noch unversehrt da.
    zwischen = ziel.with_name(f"{ziel.stem}.unfertig{ziel.suffix}")
    mappe.save(zwischen)
    os.replace(zwischen, ziel)

    return Bericht(
        datei=str(ziel),
        neu=neu,
        aktualisiert=aktualisiert,
        sicherung=sicherung,
        entfernt=entfernt,
    )
